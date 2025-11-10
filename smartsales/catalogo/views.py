import io
from decimal import Decimal, InvalidOperation
from datetime import datetime

from django.db import connection, transaction
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser, FormParser

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from smartsales.rolesusuario.permissions import IsVendedorRole
from .serializers import (
    ImportarCatalogoSerializer,
    ResultadoImportacionSerializer,
)


class DescargarPlantillaView(APIView):
    """
    Vista para descargar la plantilla de Excel para importar productos.
    Solo accesible por vendedores.
    
    GET /api/catalogo/descargar-plantilla/
    """
    permission_classes = [IsAuthenticated, IsVendedorRole]
    
    def get(self, request):
        if not OPENPYXL_AVAILABLE:
            return Response(
                {"detail": "La funcionalidad de Excel no está disponible. Instale 'openpyxl'."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        # Obtener catálogos disponibles
        with connection.cursor() as cursor:
            cursor.execute("SELECT id, nombre FROM marca ORDER BY nombre")
            marcas = cursor.fetchall()
            
            cursor.execute("SELECT id, nombre FROM tipoproducto ORDER BY nombre")
            tipos = cursor.fetchall()
        
        # Crear libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catálogo de Productos"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = [
            "NOMBRE DEL PRODUCTO*",
            "PRECIO (USD)*",
            "STOCK*",
            "TIEMPO GARANTÍA (DÍAS)*",
            "MARCA*",
            "TIPO DE PRODUCTO*"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 40  # Nombre
        ws.column_dimensions['B'].width = 15  # Precio
        ws.column_dimensions['C'].width = 12  # Stock
        ws.column_dimensions['D'].width = 20  # Garantía
        ws.column_dimensions['E'].width = 25  # Marca
        ws.column_dimensions['F'].width = 30  # Tipo
        
        # Ejemplos (primeras 3 filas)
        ejemplos = [
            ["Refrigerador Premium 500L", "899.99", "15", "365", "Samsung", "Refrigerador"],
            ["Lavadora Automática 12kg", "549.50", "8", "180", "LG", "Lavadora"],
            ["Microondas Digital 1.2 cu ft", "129.99", "25", "90", "Panasonic", "Microondas"]
        ]
        
        example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        for row_num, ejemplo in enumerate(ejemplos, 2):
            for col_num, value in enumerate(ejemplo, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.fill = example_fill
                cell.border = border
                if col_num in [2, 3, 4]:  # Precio, Stock, Garantía
                    cell.alignment = Alignment(horizontal="right")
        
        # Crear hoja de instrucciones
        ws_inst = wb.create_sheet("Instrucciones")
        ws_inst.column_dimensions['A'].width = 80
        
        instrucciones = [
            "INSTRUCCIONES PARA IMPORTAR CATÁLOGO DE PRODUCTOS",
            "",
            "1. Complete la información en la hoja 'Catálogo de Productos'",
            "2. Los campos marcados con * son OBLIGATORIOS",
            "3. No modifique los encabezados de las columnas",
            "4. Elimine las filas de ejemplo antes de importar o reemplácelas con sus datos",
            "",
            "DESCRIPCIÓN DE CAMPOS:",
            "",
            "• NOMBRE DEL PRODUCTO: Nombre descriptivo del producto (máx. 160 caracteres)",
            "• PRECIO: Precio en dólares (USD), use punto como separador decimal (ej: 99.99)",
            "• STOCK: Cantidad disponible en inventario (número entero positivo)",
            "• TIEMPO GARANTÍA: Días de garantía del producto (número entero, ej: 365 para 1 año)",
            "• MARCA: Nombre exacto de la marca. Marcas disponibles:",
        ]
        
        # Agregar marcas disponibles
        for marca in marcas:
            instrucciones.append(f"  - {marca[1]}")
        
        instrucciones.extend([
            "",
            "• TIPO DE PRODUCTO: Nombre exacto del tipo. Tipos disponibles:",
        ])
        
        # Agregar tipos disponibles
        for tipo in tipos:
            instrucciones.append(f"  - {tipo[1]}")
        
        instrucciones.extend([
            "",
            "NOTAS IMPORTANTES:",
            "",
            "• Si la marca o tipo no existe en el sistema, la fila será rechazada",
            "• El precio debe ser mayor a 0",
            "• El stock debe ser mayor o igual a 0",
            "• El tiempo de garantía debe ser mayor o igual a 0",
            "• Los productos serán asignados automáticamente a su usuario vendedor",
            "• Puede importar hasta 1000 productos a la vez",
            "",
            "EJEMPLO DE FILA VÁLIDA:",
            "Refrigerador Premium 500L | 899.99 | 15 | 365 | Samsung | Refrigerador"
        ])
        
        for row_num, linea in enumerate(instrucciones, 1):
            cell = ws_inst.cell(row=row_num, column=1, value=linea)
            if row_num == 1:
                cell.font = Font(bold=True, size=14, color="4472C4")
            elif "DESCRIPCIÓN DE CAMPOS:" in linea or "NOTAS IMPORTANTES:" in linea or "EJEMPLO DE FILA VÁLIDA:" in linea:
                cell.font = Font(bold=True, size=11)
        
        # Crear hoja de marcas
        ws_marcas = wb.create_sheet("Marcas Disponibles")
        ws_marcas['A1'] = "ID"
        ws_marcas['B1'] = "MARCA"
        ws_marcas['A1'].font = header_font
        ws_marcas['B1'].font = header_font
        ws_marcas['A1'].fill = header_fill
        ws_marcas['B1'].fill = header_fill
        
        for idx, marca in enumerate(marcas, 2):
            ws_marcas[f'A{idx}'] = marca[0]
            ws_marcas[f'B{idx}'] = marca[1]
        
        ws_marcas.column_dimensions['A'].width = 8
        ws_marcas.column_dimensions['B'].width = 30
        
        # Crear hoja de tipos
        ws_tipos = wb.create_sheet("Tipos Disponibles")
        ws_tipos['A1'] = "ID"
        ws_tipos['B1'] = "TIPO DE PRODUCTO"
        ws_tipos['A1'].font = header_font
        ws_tipos['B1'].font = header_font
        ws_tipos['A1'].fill = header_fill
        ws_tipos['B1'].fill = header_fill
        
        for idx, tipo in enumerate(tipos, 2):
            ws_tipos[f'A{idx}'] = tipo[0]
            ws_tipos[f'B{idx}'] = tipo[1]
        
        ws_tipos.column_dimensions['A'].width = 8
        ws_tipos.column_dimensions['B'].width = 35
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo con fecha
        fecha_actual = datetime.now().strftime("%Y%m%d")
        filename = f"plantilla_catalogo_productos_{fecha_actual}.xlsx"
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


class ImportarCatalogoView(APIView):
    """
    Vista para importar catálogo de productos desde archivo Excel.
    Solo accesible por vendedores.
    
    POST /api/catalogo/importar/
    """
    permission_classes = [IsAuthenticated, IsVendedorRole]
    parser_classes = [MultiPartParser, FormParser]
    
    def post(self, request):
        if not OPENPYXL_AVAILABLE:
            return Response(
                {"detail": "La funcionalidad de Excel no está disponible. Instale 'openpyxl'."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        serializer = ImportarCatalogoSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        archivo = serializer.validated_data['archivo']
        vendedor_id = request.user.id
        
        try:
            # Leer archivo Excel
            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb.active
            
            # Obtener catálogos para validación
            with connection.cursor() as cursor:
                cursor.execute("SELECT LOWER(nombre), id FROM marca")
                marcas_map = {row[0]: row[1] for row in cursor.fetchall()}
                
                cursor.execute("SELECT LOWER(nombre), id FROM tipoproducto")
                tipos_map = {row[0]: row[1] for row in cursor.fetchall()}
            
            errores = []
            productos_creados = []
            total_procesados = 0
            exitosos = 0
            fallidos = 0
            
            # Procesar filas (empezar desde la fila 2, después de encabezados)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=1002, values_only=True), start=2):
                # Saltar filas vacías
                if not any(row):
                    continue
                
                total_procesados += 1
                
                # Limitar a 1000 productos
                if total_procesados > 1000:
                    errores.append({
                        'fila': row_num,
                        'error': 'Se alcanzó el límite de 1000 productos por importación'
                    })
                    break
                
                try:
                    # Extraer valores
                    nombre = str(row[0]).strip() if row[0] else None
                    precio = row[1]
                    stock = row[2]
                    tiempo_garantia = row[3]
                    marca_nombre = str(row[4]).strip() if row[4] else None
                    tipo_nombre = str(row[5]).strip() if row[5] else None
                    
                    # Validar campos obligatorios
                    if not nombre:
                        raise ValueError("El nombre del producto es obligatorio")
                    if len(nombre) > 160:
                        raise ValueError("El nombre no debe superar 160 caracteres")
                    
                    if precio is None:
                        raise ValueError("El precio es obligatorio")
                    try:
                        precio = Decimal(str(precio))
                        if precio <= 0:
                            raise ValueError("El precio debe ser mayor a 0")
                    except (InvalidOperation, ValueError) as e:
                        raise ValueError(f"Precio inválido: {precio}")
                    
                    if stock is None:
                        raise ValueError("El stock es obligatorio")
                    try:
                        stock = int(stock)
                        if stock < 0:
                            raise ValueError("El stock no puede ser negativo")
                    except (ValueError, TypeError):
                        raise ValueError(f"Stock inválido: {stock}")
                    
                    if tiempo_garantia is None:
                        raise ValueError("El tiempo de garantía es obligatorio")
                    try:
                        tiempo_garantia = int(tiempo_garantia)
                        if tiempo_garantia < 0:
                            raise ValueError("El tiempo de garantía no puede ser negativo")
                    except (ValueError, TypeError):
                        raise ValueError(f"Tiempo de garantía inválido: {tiempo_garantia}")
                    
                    if not marca_nombre:
                        raise ValueError("La marca es obligatoria")
                    
                    marca_id = marcas_map.get(marca_nombre.lower())
                    if not marca_id:
                        raise ValueError(f"Marca '{marca_nombre}' no encontrada. Revise la hoja 'Marcas Disponibles'")
                    
                    if not tipo_nombre:
                        raise ValueError("El tipo de producto es obligatorio")
                    
                    tipo_id = tipos_map.get(tipo_nombre.lower())
                    if not tipo_id:
                        raise ValueError(f"Tipo '{tipo_nombre}' no encontrado. Revise la hoja 'Tipos Disponibles'")
                    
                    # Insertar producto
                    with transaction.atomic():
                        with connection.cursor() as cursor:
                            cursor.execute(
                                """
                                INSERT INTO producto 
                                (nombre, precio, stock, tiempogarantia, marca_id, tipoproducto_id, id_vendedor)
                                VALUES (%s, %s, %s, %s, %s, %s, %s)
                                RETURNING id
                                """,
                                [nombre, precio, stock, tiempo_garantia, marca_id, tipo_id, vendedor_id]
                            )
                            producto_id = cursor.fetchone()[0]
                    
                    productos_creados.append({
                        'id': producto_id,
                        'nombre': nombre,
                        'precio': float(precio),
                        'stock': stock,
                        'fila': row_num
                    })
                    exitosos += 1
                    
                except Exception as e:
                    errores.append({
                        'fila': row_num,
                        'error': str(e),
                        'datos': {
                            'nombre': nombre if 'nombre' in locals() else None,
                            'precio': str(precio) if 'precio' in locals() else None,
                        }
                    })
                    fallidos += 1
            
            resultado = {
                'total_procesados': total_procesados,
                'exitosos': exitosos,
                'fallidos': fallidos,
                'errores': errores,
                'productos_creados': productos_creados
            }
            
            return Response(
                ResultadoImportacionSerializer(resultado).data,
                status=status.HTTP_201_CREATED if exitosos > 0 else status.HTTP_400_BAD_REQUEST
            )
            
        except Exception as e:
            return Response(
                {"detail": f"Error al procesar el archivo: {str(e)}"},
                status=status.HTTP_400_BAD_REQUEST
            )


class ExportarCatalogoView(APIView):
    """
    Vista para exportar el catálogo de productos del vendedor a Excel.
    Solo accesible por vendedores.
    
    GET /api/catalogo/exportar/
    """
    permission_classes = [IsAuthenticated, IsVendedorRole]
    
    def get(self, request):
        if not OPENPYXL_AVAILABLE:
            return Response(
                {"detail": "La funcionalidad de Excel no está disponible. Instale 'openpyxl'."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        vendedor_id = request.user.id
        vendedor_nombre = request.user.nombre if hasattr(request.user, 'nombre') else 'Vendedor'
        
        # Obtener productos del vendedor
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT 
                    p.id,
                    p.nombre,
                    p.precio,
                    p.stock,
                    p.tiempogarantia,
                    m.nombre as marca,
                    tp.nombre as tipo
                FROM producto p
                JOIN marca m ON p.marca_id = m.id
                JOIN tipoproducto tp ON p.tipoproducto_id = tp.id
                WHERE p.id_vendedor = %s
                ORDER BY p.nombre
                """,
                [vendedor_id]
            )
            productos = cursor.fetchall()
        
        if not productos:
            return Response(
                {"detail": "No tienes productos en tu catálogo para exportar."},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Crear libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mi Catálogo"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        data_alignment_center = Alignment(horizontal="center")
        data_alignment_right = Alignment(horizontal="right")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f"CATÁLOGO DE PRODUCTOS - {vendedor_nombre.upper()}"
        title_cell.font = Font(bold=True, size=14, color="4472C4")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Fecha de exportación
        ws.merge_cells('A2:G2')
        date_cell = ws['A2']
        date_cell.value = f"Exportado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        date_cell.alignment = Alignment(horizontal="center")
        date_cell.font = Font(italic=True, size=10)
        
        # Encabezados (fila 4)
        headers = ["ID", "NOMBRE", "PRECIO (USD)", "STOCK", "GARANTÍA (DÍAS)", "MARCA", "TIPO"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Datos
        for row_num, producto in enumerate(productos, 5):
            ws.cell(row=row_num, column=1, value=producto[0]).alignment = data_alignment_center
            ws.cell(row=row_num, column=2, value=producto[1])
            ws.cell(row=row_num, column=3, value=float(producto[2])).number_format = '$#,##0.00'
            ws.cell(row=row_num, column=3).alignment = data_alignment_right
            ws.cell(row=row_num, column=4, value=producto[3]).alignment = data_alignment_center
            ws.cell(row=row_num, column=5, value=producto[4]).alignment = data_alignment_center
            ws.cell(row=row_num, column=6, value=producto[5])
            ws.cell(row=row_num, column=7, value=producto[6])
            
            # Aplicar bordes
            for col in range(1, 8):
                ws.cell(row=row_num, column=col).border = border
        
        # Totales (última fila)
        total_row = len(productos) + 5
        ws.merge_cells(f'A{total_row}:B{total_row}')
        total_cell = ws[f'A{total_row}']
        total_cell.value = f"TOTAL PRODUCTOS: {len(productos)}"
        total_cell.font = Font(bold=True)
        total_cell.alignment = Alignment(horizontal="right")
        
        # Calcular valor total del inventario
        valor_total = sum(float(p[2]) * p[3] for p in productos)
        ws[f'C{total_row}'] = valor_total
        ws[f'C{total_row}'].number_format = '$#,##0.00'
        ws[f'C{total_row}'].font = Font(bold=True)
        ws[f'C{total_row}'].alignment = data_alignment_right
        
        total_stock = sum(p[3] for p in productos)
        ws[f'D{total_row}'] = total_stock
        ws[f'D{total_row}'].font = Font(bold=True)
        ws[f'D{total_row}'].alignment = data_alignment_center
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 25
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre de archivo
        fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"catalogo_{vendedor_nombre.replace(' ', '_')}_{fecha_actual}.xlsx"
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
